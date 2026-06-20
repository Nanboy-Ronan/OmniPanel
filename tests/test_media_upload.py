"""
tests/test_media_upload.py  —  TDD tests for WeChat Official Account manual upload.

Written BEFORE the implementation.  All tests should initially fail (ImportError
or 404/422) and pass once the feature is built.

Test groups
-----------
1.  Parser unit tests (pure Python, no DB)
      app.db.media_etl.parse_wechat_xlsx(filepath, account_id)
      → (rows: list[dict], rejected: list[str])

2.  MediaSyncRun model — new columns (source, filename)

3.  Account management API
      POST  /media/accounts          admin only

4.  Upload endpoint
      POST  /media/upload            multipart: file + account_id form field
      analysts and admins allowed; viewers forbidden

5.  Upload history
      GET   /media/uploads           analyst+
"""
from __future__ import annotations

import hashlib
import io
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pytest

# DISABLED 2026-06-01: xlsx upload was never used in production; all tests skipped.
# Remove this mark to re-enable. See README §"Disabled: WeChat xlsx upload".
pytestmark = pytest.mark.skip(reason="xlsx upload disabled — data comes from API sync, see README")

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

DATA_DIR = PROJECT_ROOT / "data"
REAL_XLSX = DATA_DIR / "pantaobencao.xlsx"

# ── Re-use client / tokens / _auth from the canonical fixture module ────────
from tests.test_api_endpoints import _auth, client, tokens  # noqa: F401


# ═══════════════════════════════════════════════════════════════════════════════
# Fixture helpers
# ═══════════════════════════════════════════════════════════════════════════════

# Columns present in a real WeChat backend export
_WECHAT_HEADERS = [
    "文章名称", "发布日期", "发布时间",
    "阅读人数", "点赞人数", "分享人数", "推荐人数", "留言条数", "划线人数",
    "被转载次数", "阅读（次）", "平均阅读时长（分钟）", "完读率", "阅读后关注",
    "听全文", "推送",
]


def _make_wechat_xlsx_bytes(rows: list[list] | None = None) -> bytes:
    """Build a minimal valid WeChat export xlsx in memory via openpyxl.

    ``rows`` is a list of row-value lists (excluding the header row).
    Default: two articles with different date string formats.
    """
    if rows is None:
        rows = [
            # title,           date,           time,   reads, likes, shares, rec, comments, mark,
            # reprints, read_count, avg_time, completion, follow, listen, push
            ["健康早报一", "2026/5/22",          "19:01", 200, 5, 10, 1, 0, 3,  0, 210, 2.5, 0.45, 1, 0, "Y"],
            ["健康早报二", "2026/5/21",          "18:00", 150, 3,  7, 0, 1, 1,  0, 155, 1.8, 0.38, 0, 0, "Y"],
            ["健康早报三", "2025-04-25 00:00:00", "17:01",  80, 2,  4, 2, 0, 0,  0,  85, 0.9, 0.30, 2, 0, "Y"],
        ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工作表1"
    ws.append(_WECHAT_HEADERS)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_wechat_xlsx_multi_sheet_bytes() -> bytes:
    """Two data sheets + one formula-definition sheet (no 文章名称 column)."""
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "工作表1"
    ws1.append(_WECHAT_HEADERS)
    ws1.append(["文章甲", "2026/5/10", "19:00", 100, 2, 5, 0, 0, 0, 0, 100, 1.0, 0.4, 0, 0, "Y"])

    ws2 = wb.create_sheet("工作表1(副本)")
    ws2.append(_WECHAT_HEADERS)
    ws2.append(["文章乙(旧版)", "2025/4/1", "10:00", 60, 1, 3, 0, 0, 0, 0, 60, 0.8, 0.3, 0, 0, "Y"])

    ws3 = wb.create_sheet("工作表2")
    ws3.append(["指标", "公式", "含义"])
    ws3.append(["点赞率", "点赞人数÷阅读人数", "基础认可"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _create_account(client, admin_token: str, name: str = "示例公众号") -> dict:
    """POST /media/accounts and return the JSON body."""
    r = client.post(
        "/media/accounts",
        json={"name": name},
        headers=_auth(admin_token),
    )
    assert r.status_code == 201, f"create account failed: {r.text}"
    return r.json()


# ═══════════════════════════════════════════════════════════════════════════════
# Section 1 — Parser unit tests (no DB required)
# ═══════════════════════════════════════════════════════════════════════════════

class TestWeChatXlsxParser:
    """Tests for app.db.media_etl.parse_wechat_xlsx()."""

    # ------------------------------------------------------------------ import
    @pytest.fixture(autouse=True)
    def _import_parser(self):
        from app.db.media_etl import parse_wechat_xlsx
        self.parse = parse_wechat_xlsx

    # ------------------------------------------------------------------ basics
    def test_parse_returns_list_of_dicts(self, tmp_path):
        xlsx = tmp_path / "export.xlsx"
        xlsx.write_bytes(_make_wechat_xlsx_bytes())
        rows, rejected = self.parse(xlsx, account_id=1)
        assert isinstance(rows, list)
        assert isinstance(rejected, list)

    def test_parse_correct_row_count(self, tmp_path):
        xlsx = tmp_path / "export.xlsx"
        xlsx.write_bytes(_make_wechat_xlsx_bytes())
        rows, rejected = self.parse(xlsx, account_id=1)
        assert len(rows) == 3
        assert len(rejected) == 0

    def test_parse_row_has_required_keys(self, tmp_path):
        xlsx = tmp_path / "export.xlsx"
        xlsx.write_bytes(_make_wechat_xlsx_bytes())
        rows, _ = self.parse(xlsx, account_id=1)
        required = {
            "external_id", "title", "publish_date",
            "read_user_count", "read_count",
            "like_user", "share_user_count",
            "comment_count", "collection_user",
            "read_avg_time", "raw_payload",
        }
        for row in rows:
            assert required.issubset(row.keys()), f"Missing keys: {required - row.keys()}"

    # ------------------------------------------------------------------ dates
    def test_parse_date_slash_format(self, tmp_path):
        """'2026/5/22' → date(2026, 5, 22)"""
        xlsx = tmp_path / "export.xlsx"
        xlsx.write_bytes(_make_wechat_xlsx_bytes())
        rows, _ = self.parse(xlsx, account_id=1)
        assert rows[0]["publish_date"] == date(2026, 5, 22)

    def test_parse_date_iso_format(self, tmp_path):
        """'2025-04-25 00:00:00' → date(2025, 4, 25)"""
        xlsx = tmp_path / "export.xlsx"
        xlsx.write_bytes(_make_wechat_xlsx_bytes())
        rows, _ = self.parse(xlsx, account_id=1)
        assert rows[2]["publish_date"] == date(2025, 4, 25)

    def test_parse_date_slash_with_time_suffix(self, tmp_path):
        """'2025/5/3-17:06' → date(2025, 5, 3)"""
        single_row = [
            ["文章测试", "2025/5/3-17:06", "17:06",
             50, 1, 2, 0, 0, 0, 0, 52, 1.1, 0.3, 0, 0, "Y"],
        ]
        xlsx = tmp_path / "export.xlsx"
        xlsx.write_bytes(_make_wechat_xlsx_bytes(rows=single_row))
        rows, _ = self.parse(xlsx, account_id=1)
        assert rows[0]["publish_date"] == date(2025, 5, 3)

    def test_parse_datetime_object_in_cell(self, tmp_path):
        """openpyxl may produce datetime objects in cells; parser must handle both."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "工作表1"
        ws.append(_WECHAT_HEADERS)
        ws.append([
            "文章datetime",
            datetime(2026, 3, 15, 19, 0, 0),  # cell contains a Python datetime
            "19:00", 77, 2, 3, 0, 0, 0, 0, 77, 1.2, 0.35, 0, 0, "Y",
        ])
        buf = io.BytesIO(); wb.save(buf)
        xlsx = tmp_path / "export.xlsx"
        xlsx.write_bytes(buf.getvalue())
        rows, _ = self.parse(xlsx, account_id=1)
        assert rows[0]["publish_date"] == date(2026, 3, 15)

    # ------------------------------------------------------------------ numeric fields
    def test_parse_numeric_fields(self, tmp_path):
        xlsx = tmp_path / "export.xlsx"
        xlsx.write_bytes(_make_wechat_xlsx_bytes())
        rows, _ = self.parse(xlsx, account_id=1)
        r = rows[0]
        assert r["read_user_count"] == 200
        assert r["read_count"] == 210
        assert r["like_user"] == 5
        assert r["share_user_count"] == 10
        assert r["comment_count"] == 0
        assert r["collection_user"] == 3
        assert abs(r["read_avg_time"] - 2.5) < 1e-9

    def test_parse_avg_time_none_when_missing(self, tmp_path):
        rows_data = [
            ["文章无时长", "2026/5/1", "19:00",
             50, 1, 2, 0, 0, 0, 0, 50, None, 0.3, 0, 0, "Y"],
        ]
        xlsx = tmp_path / "export.xlsx"
        xlsx.write_bytes(_make_wechat_xlsx_bytes(rows=rows_data))
        rows, _ = self.parse(xlsx, account_id=1)
        assert rows[0]["read_avg_time"] is None

    def test_parse_string_numerics_coerced(self, tmp_path):
        """WeChat sometimes exports numbers as strings; parser should coerce."""
        rows_data = [
            ["文章字符串数字", "2026/5/1", "19:00",
             "300", "8", "15", "2", "1", "4", "0", "305", "3.0", "0.5", "2", "0", "Y"],
        ]
        xlsx = tmp_path / "export.xlsx"
        xlsx.write_bytes(_make_wechat_xlsx_bytes(rows=rows_data))
        rows, _ = self.parse(xlsx, account_id=1)
        assert rows[0]["read_user_count"] == 300
        assert rows[0]["like_user"] == 8

    # ------------------------------------------------------------------ external_id
    def test_parse_external_id_is_deterministic(self, tmp_path):
        """Same account_id + title always produces the same external_id."""
        xlsx = tmp_path / "export.xlsx"
        xlsx.write_bytes(_make_wechat_xlsx_bytes())
        rows_a, _ = self.parse(xlsx, account_id=1)
        rows_b, _ = self.parse(xlsx, account_id=1)
        assert rows_a[0]["external_id"] == rows_b[0]["external_id"]

    def test_parse_external_id_differs_by_account(self, tmp_path):
        """Same file with different account_id must produce different external_ids."""
        xlsx = tmp_path / "export.xlsx"
        xlsx.write_bytes(_make_wechat_xlsx_bytes())
        rows_a, _ = self.parse(xlsx, account_id=1)
        rows_b, _ = self.parse(xlsx, account_id=99)
        assert rows_a[0]["external_id"] != rows_b[0]["external_id"]

    def test_parse_external_id_is_sha256_prefix(self, tmp_path):
        """external_id should be the first 32 hex chars of sha256(account_id:title)."""
        xlsx = tmp_path / "export.xlsx"
        xlsx.write_bytes(_make_wechat_xlsx_bytes())
        rows, _ = self.parse(xlsx, account_id=7)
        title = rows[0]["title"]
        expected = hashlib.sha256(f"7:{title}".encode()).hexdigest()[:32]
        assert rows[0]["external_id"] == expected

    # ------------------------------------------------------------------ multi-sheet
    def test_parse_picks_first_data_sheet(self, tmp_path):
        """With multiple sheets, parser uses the first sheet that has 文章名称 column."""
        xlsx = tmp_path / "multi.xlsx"
        xlsx.write_bytes(_make_wechat_xlsx_multi_sheet_bytes())
        rows, _ = self.parse(xlsx, account_id=1)
        # Only sheet1 data (文章甲); sheet2 (文章乙) and sheet3 are skipped
        assert len(rows) == 1
        assert rows[0]["title"] == "文章甲"

    def test_parse_skips_sheet_without_title_column(self, tmp_path):
        """A sheet without 文章名称 column is silently ignored."""
        wb = openpyxl.Workbook()
        ws_bad = wb.active
        ws_bad.title = "公式"
        ws_bad.append(["指标", "公式", "含义"])
        ws_bad.append(["点赞率", "点赞/阅读", "认可"])

        ws_good = wb.create_sheet("数据")
        ws_good.append(_WECHAT_HEADERS)
        ws_good.append(["隐藏文章", "2026/4/1", "19:00",
                         88, 2, 4, 0, 0, 0, 0, 90, 1.5, 0.4, 0, 0, "Y"])
        buf = io.BytesIO(); wb.save(buf)
        xlsx = tmp_path / "bad_first.xlsx"
        xlsx.write_bytes(buf.getvalue())
        rows, _ = self.parse(xlsx, account_id=1)
        assert len(rows) == 1
        assert rows[0]["title"] == "隐藏文章"

    # ------------------------------------------------------------------ rejected rows
    def test_parse_rejects_row_with_empty_title(self, tmp_path):
        """Rows with empty 文章名称 go to the rejected list."""
        rows_data = [
            ["有效文章",  "2026/5/1", "19:00", 100, 3, 5, 0, 0, 0, 0, 100, 1.0, 0.4, 0, 0, "Y"],
            ["",          "2026/5/1", "19:00",  10, 0, 0, 0, 0, 0, 0,  10, 0.5, 0.1, 0, 0, "Y"],
            [None,        "2026/5/2", "08:00",  20, 1, 2, 0, 0, 0, 0,  20, 0.8, 0.2, 0, 0, "Y"],
        ]
        xlsx = tmp_path / "export.xlsx"
        xlsx.write_bytes(_make_wechat_xlsx_bytes(rows=rows_data))
        rows, rejected = self.parse(xlsx, account_id=1)
        assert len(rows) == 1
        assert len(rejected) == 2

    def test_parse_rejects_row_with_unparseable_date(self, tmp_path):
        rows_data = [
            ["正常文章",   "2026/5/1",   "19:00", 50, 1, 2, 0, 0, 0, 0, 50, 1.0, 0.3, 0, 0, "Y"],
            ["日期损坏",   "INVALID",    "19:00", 30, 0, 1, 0, 0, 0, 0, 30, 0.5, 0.2, 0, 0, "Y"],
        ]
        xlsx = tmp_path / "export.xlsx"
        xlsx.write_bytes(_make_wechat_xlsx_bytes(rows=rows_data))
        rows, rejected = self.parse(xlsx, account_id=1)
        assert len(rows) == 1
        assert len(rejected) == 1

    # ------------------------------------------------------------------ raw_payload
    def test_parse_raw_payload_captures_extra_columns(self, tmp_path):
        """Columns not mapped to dedicated fields end up in raw_payload."""
        xlsx = tmp_path / "export.xlsx"
        xlsx.write_bytes(_make_wechat_xlsx_bytes())
        rows, _ = self.parse(xlsx, account_id=1)
        # 推荐人数 (rec) has no dedicated DB column → must be in raw_payload
        assert "推荐人数" in rows[0]["raw_payload"]

    # ------------------------------------------------------------------ real file
    @pytest.mark.skipif(not REAL_XLSX.exists(), reason="pantaobencao.xlsx not in data/")
    def test_parse_real_pantaobencao_file(self):
        """Smoke-test against the actual sample file — must not raise."""
        rows, rejected = self.parse(REAL_XLSX, account_id=1)
        # The real file has 167 + 92 rows across two data sheets; first sheet wins
        assert len(rows) > 0
        # Every row should have a title
        for r in rows:
            assert r["title"]
            assert r["publish_date"] is not None
            assert isinstance(r["read_user_count"], int)


# ═══════════════════════════════════════════════════════════════════════════════
# Section 2 — MediaSyncRun model: new columns
# ═══════════════════════════════════════════════════════════════════════════════

class TestMediaSyncRunNewColumns:

    def test_sync_run_has_source_column(self, pg_sync_url):
        from sqlalchemy import create_engine, inspect
        engine = create_engine(pg_sync_url, future=True)
        try:
            inspector = inspect(engine)
            cols = {c["name"] for c in inspector.get_columns("media_sync_runs")}
        finally:
            engine.dispose()
        assert "source" in cols, (
            "'source' column missing from media_sync_runs — "
            "add it in models.py + ADDITIVE_DDL"
        )

    def test_sync_run_has_filename_column(self, pg_sync_url):
        from sqlalchemy import create_engine, inspect
        engine = create_engine(pg_sync_url, future=True)
        try:
            inspector = inspect(engine)
            cols = {c["name"] for c in inspector.get_columns("media_sync_runs")}
        finally:
            engine.dispose()
        assert "filename" in cols, (
            "'filename' column missing from media_sync_runs — "
            "add it in models.py + ADDITIVE_DDL"
        )

    def test_sync_run_source_defaults_to_api(self, pg_sync_url):
        """Existing rows created without source should default to 'api'."""
        from sqlalchemy import create_engine, text
        engine = create_engine(pg_sync_url, future=True)
        try:
            with engine.begin() as conn:
                # Insert a run without the source column to test the default
                conn.execute(text("""
                    INSERT INTO media_accounts (platform, name, is_active)
                    VALUES ('wechat_official', 'DefaultTest', TRUE)
                """))
                acc_id = conn.execute(
                    text("SELECT id FROM media_accounts WHERE name='DefaultTest'")
                ).scalar()
                conn.execute(text("""
                    INSERT INTO media_sync_runs
                        (account_id, status, start_date, end_date)
                    VALUES (:aid, 'success', '2026-01-01', '2026-01-01')
                """), {"aid": acc_id})
                source = conn.execute(text(
                    "SELECT source FROM media_sync_runs WHERE account_id=:aid"
                ), {"aid": acc_id}).scalar()
            assert source == "api"
        finally:
            engine.dispose()

    def test_sync_run_source_can_be_manual(self, pg_sync_url):
        from sqlalchemy import create_engine
        from sqlalchemy.orm import sessionmaker
        from app.db.models import MediaAccount, MediaSyncRun
        engine = create_engine(pg_sync_url, future=True)
        Session = sessionmaker(bind=engine)
        try:
            with Session() as s:
                acc = MediaAccount(
                    platform="wechat_official", name="ManualTest", is_active=True
                )
                s.add(acc)
                s.flush()
                run = MediaSyncRun(
                    account_id=acc.id,
                    status="success",
                    start_date=date(2026, 5, 1),
                    end_date=date(2026, 5, 1),
                    source="manual",
                    filename="pantaobencao.xlsx",
                )
                s.add(run)
                s.commit()

                stored = s.query(MediaSyncRun).filter_by(
                    account_id=acc.id, source="manual"
                ).one()
                assert stored.source == "manual"
                assert stored.filename == "pantaobencao.xlsx"
        finally:
            engine.dispose()


# ═══════════════════════════════════════════════════════════════════════════════
# Section 3 — Account management  POST /media/accounts
# ═══════════════════════════════════════════════════════════════════════════════

class TestMediaAccountCreation:

    def test_admin_can_create_account(self, client, tokens):
        r = client.post(
            "/media/accounts",
            json={"name": "示例公众号"},
            headers=_auth(tokens["admin"]),
        )
        assert r.status_code == 201
        body = r.json()
        assert body["name"] == "示例公众号"
        assert body["platform"] == "wechat_official"
        assert "id" in body

    def test_analyst_cannot_create_account(self, client, tokens):
        r = client.post(
            "/media/accounts",
            json={"name": "未授权账号"},
            headers=_auth(tokens["analyst"]),
        )
        assert r.status_code == 403

    def test_viewer_cannot_create_account(self, client, tokens):
        r = client.post(
            "/media/accounts",
            json={"name": "未授权账号"},
            headers=_auth(tokens["viewer"]),
        )
        assert r.status_code == 403

    def test_unauthenticated_cannot_create_account(self, client):
        r = client.post("/media/accounts", json={"name": "无token账号"})
        assert r.status_code in (401, 403)

    def test_create_account_app_id_is_optional(self, client, tokens):
        """Manual accounts have no app_id; it must be nullable."""
        r = client.post(
            "/media/accounts",
            json={"name": "无AppID账号"},
            headers=_auth(tokens["admin"]),
        )
        assert r.status_code == 201
        assert r.json().get("app_id") is None

    def test_create_account_with_explicit_app_id(self, client, tokens):
        r = client.post(
            "/media/accounts",
            json={"name": "有AppID账号", "app_id": "wx-test-123"},
            headers=_auth(tokens["admin"]),
        )
        assert r.status_code == 201
        assert r.json()["app_id"] == "wx-test-123"

    def test_created_account_appears_in_list(self, client, tokens):
        client.post(
            "/media/accounts",
            json={"name": "列表测试账号"},
            headers=_auth(tokens["admin"]),
        )
        r = client.get("/media/accounts", headers=_auth(tokens["analyst"]))
        assert r.status_code == 200
        names = [a["name"] for a in r.json()]
        assert "列表测试账号" in names

    def test_missing_name_returns_422(self, client, tokens):
        r = client.post(
            "/media/accounts",
            json={"app_id": "wx-no-name"},
            headers=_auth(tokens["admin"]),
        )
        assert r.status_code == 422


# ═══════════════════════════════════════════════════════════════════════════════
# Section 4 — Upload endpoint  POST /media/upload
# ═══════════════════════════════════════════════════════════════════════════════

class TestMediaUploadEndpoint:

    # ------------------------------------------------------------------ auth
    def test_upload_requires_auth(self, client, tokens):
        xlsx_bytes = _make_wechat_xlsx_bytes()
        r = client.post(
            "/media/upload",
            files={"file": ("export.xlsx", xlsx_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"account_id": "1"},
        )
        assert r.status_code in (401, 403)

    def test_viewer_cannot_upload(self, client, tokens):
        acc = _create_account(client, tokens["admin"])
        xlsx_bytes = _make_wechat_xlsx_bytes()
        r = client.post(
            "/media/upload",
            files={"file": ("export.xlsx", xlsx_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"account_id": str(acc["id"])},
            headers=_auth(tokens["viewer"]),
        )
        assert r.status_code == 403

    def test_analyst_can_upload(self, client, tokens):
        acc = _create_account(client, tokens["admin"])
        xlsx_bytes = _make_wechat_xlsx_bytes()
        r = client.post(
            "/media/upload",
            files={"file": ("export.xlsx", xlsx_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"account_id": str(acc["id"])},
            headers=_auth(tokens["analyst"]),
        )
        assert r.status_code == 200, r.text

    # ------------------------------------------------------------------ validation
    def test_upload_without_account_id_returns_422(self, client, tokens):
        xlsx_bytes = _make_wechat_xlsx_bytes()
        r = client.post(
            "/media/upload",
            files={"file": ("export.xlsx", xlsx_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=_auth(tokens["admin"]),
            # no data={"account_id": ...}
        )
        assert r.status_code == 422

    def test_upload_invalid_account_id_returns_404(self, client, tokens):
        xlsx_bytes = _make_wechat_xlsx_bytes()
        r = client.post(
            "/media/upload",
            files={"file": ("export.xlsx", xlsx_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"account_id": "99999"},
            headers=_auth(tokens["admin"]),
        )
        assert r.status_code == 404

    def test_upload_non_xlsx_file_rejected(self, client, tokens):
        acc = _create_account(client, tokens["admin"])
        r = client.post(
            "/media/upload",
            files={"file": ("export.py", b"print('hello')", "text/plain")},
            data={"account_id": str(acc["id"])},
            headers=_auth(tokens["admin"]),
        )
        assert r.status_code == 400

    def test_upload_oversized_file_rejected(self, client, tokens):
        acc = _create_account(client, tokens["admin"])
        import os
        max_mb = int(os.getenv("MAX_UPLOAD_MB", "50"))
        big = b"x" * (max_mb * 1024 * 1024 + 1)
        r = client.post(
            "/media/upload",
            files={"file": ("big.xlsx", big,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"account_id": str(acc["id"])},
            headers=_auth(tokens["admin"]),
        )
        assert r.status_code == 413

    # ------------------------------------------------------------------ happy path
    def test_upload_returns_summary_json(self, client, tokens):
        acc = _create_account(client, tokens["admin"])
        xlsx_bytes = _make_wechat_xlsx_bytes()
        r = client.post(
            "/media/upload",
            files={"file": ("export.xlsx", xlsx_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"account_id": str(acc["id"])},
            headers=_auth(tokens["admin"]),
        )
        assert r.status_code == 200
        body = r.json()
        assert "posts_upserted" in body
        # Manual uploads no longer write to media_post_metrics_daily;
        # metrics_upserted is always 0 (flat table, no daily breakdown).
        assert "metrics_upserted" in body
        assert body["metrics_upserted"] == 0
        assert "rejected" in body

    def test_upload_inserts_correct_post_count(self, client, tokens):
        acc = _create_account(client, tokens["admin"])
        xlsx_bytes = _make_wechat_xlsx_bytes()  # 3 articles
        r = client.post(
            "/media/upload",
            files={"file": ("export.xlsx", xlsx_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"account_id": str(acc["id"])},
            headers=_auth(tokens["admin"]),
        )
        assert r.status_code == 200
        body = r.json()
        assert body["posts_upserted"] == 3   # inserted + updated
        assert body["inserted"] == 3          # all new on first upload
        assert body["updated"] == 0
        assert body["rejected"] == 0

    def test_upload_articles_queryable_via_traffic_endpoint(self, client, tokens):
        """Manual uploads must be queryable via GET /media/traffic (not /media/posts)."""
        acc = _create_account(client, tokens["admin"])
        xlsx_bytes = _make_wechat_xlsx_bytes()
        client.post(
            "/media/upload",
            files={"file": ("export.xlsx", xlsx_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"account_id": str(acc["id"])},
            headers=_auth(tokens["admin"]),
        )
        r = client.get(
            "/media/traffic",
            params={"account_id": acc["id"]},
            headers=_auth(tokens["analyst"]),
        )
        assert r.status_code == 200
        titles = [p["title"] for p in r.json()]
        assert "健康早报一" in titles
        assert "健康早报二" in titles
        assert "健康早报三" in titles

    def test_upload_metrics_values_match_xlsx(self, client, tokens):
        acc = _create_account(client, tokens["admin"])
        xlsx_bytes = _make_wechat_xlsx_bytes()
        client.post(
            "/media/upload",
            files={"file": ("export.xlsx", xlsx_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"account_id": str(acc["id"])},
            headers=_auth(tokens["admin"]),
        )
        r = client.get(
            "/media/traffic",
            params={"account_id": acc["id"]},
            headers=_auth(tokens["analyst"]),
        )
        articles = {p["title"]: p for p in r.json()}
        p = articles["健康早报一"]
        assert p["read_user_count"] == 200
        assert p["like_user"] == 5
        assert p["share_user_count"] == 10
        assert p["comment_count"] == 0
        assert p["collection_user"] == 3

    def test_upload_publish_date_stored_correctly(self, client, tokens):
        """publish_date in the traffic record should match the xlsx value."""
        acc = _create_account(client, tokens["admin"])
        xlsx_bytes = _make_wechat_xlsx_bytes()
        client.post(
            "/media/upload",
            files={"file": ("export.xlsx", xlsx_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"account_id": str(acc["id"])},
            headers=_auth(tokens["admin"]),
        )
        r = client.get(
            "/media/traffic",
            params={"account_id": acc["id"]},
            headers=_auth(tokens["analyst"]),
        )
        articles = {p["title"]: p for p in r.json()}
        assert articles["健康早报一"]["publish_date"] == "2026-05-22"

    # ------------------------------------------------------------------ idempotency
    def test_upload_twice_same_file_no_duplicates(self, client, tokens):
        """Uploading the identical file twice must not duplicate rows in media_article_traffic."""
        acc = _create_account(client, tokens["admin"])
        xlsx_bytes = _make_wechat_xlsx_bytes()

        for _ in range(2):
            client.post(
                "/media/upload",
                files={"file": ("export.xlsx", xlsx_bytes,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                data={"account_id": str(acc["id"])},
                headers=_auth(tokens["admin"]),
            )

        r = client.get(
            "/media/traffic",
            params={"account_id": acc["id"]},
            headers=_auth(tokens["analyst"]),
        )
        # Should still be exactly 3 articles — not 6
        assert len(r.json()) == 3

    def test_upload_second_time_updates_metrics(self, client, tokens):
        """A re-upload with updated numbers must overwrite the stored traffic row."""
        acc = _create_account(client, tokens["admin"])
        # First upload: 健康早报一 has 200 readers
        first_bytes = _make_wechat_xlsx_bytes()
        client.post(
            "/media/upload",
            files={"file": ("export.xlsx", first_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"account_id": str(acc["id"])},
            headers=_auth(tokens["admin"]),
        )

        # Second upload: same article now has 350 readers
        updated_rows = [
            ["健康早报一", "2026/5/22", "19:01", 350, 8, 14, 1, 0, 5, 0, 360, 2.8, 0.50, 2, 0, "Y"],
        ]
        second_bytes = _make_wechat_xlsx_bytes(rows=updated_rows)
        r2 = client.post(
            "/media/upload",
            files={"file": ("export_v2.xlsx", second_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"account_id": str(acc["id"])},
            headers=_auth(tokens["admin"]),
        )
        assert r2.json()["updated"] == 1   # existing row was updated, not inserted

        r = client.get(
            "/media/traffic",
            params={"account_id": acc["id"]},
            headers=_auth(tokens["analyst"]),
        )
        articles = {p["title"]: p for p in r.json()}
        # Must reflect the updated value, not the original 200
        assert articles["健康早报一"]["read_user_count"] == 350

    # ------------------------------------------------------------------ sync run tracking
    def test_upload_creates_sync_run_with_source_manual(self, client, tokens):
        acc = _create_account(client, tokens["admin"])
        xlsx_bytes = _make_wechat_xlsx_bytes()
        client.post(
            "/media/upload",
            files={"file": ("export.xlsx", xlsx_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"account_id": str(acc["id"])},
            headers=_auth(tokens["admin"]),
        )
        r = client.get(
            "/media/uploads",
            headers=_auth(tokens["analyst"]),
        )
        assert r.status_code == 200
        runs = r.json()
        assert len(runs) >= 1
        # Only manual uploads appear here
        assert all(run["source"] == "manual" for run in runs)

    def test_upload_filename_stored_in_sync_run(self, client, tokens):
        acc = _create_account(client, tokens["admin"])
        xlsx_bytes = _make_wechat_xlsx_bytes()
        client.post(
            "/media/upload",
            files={"file": ("pantaobencao_2026.xlsx", xlsx_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"account_id": str(acc["id"])},
            headers=_auth(tokens["admin"]),
        )
        r = client.get("/media/uploads", headers=_auth(tokens["analyst"]))
        runs = r.json()
        assert runs[0]["filename"] == "pantaobencao_2026.xlsx"

    # ------------------------------------------------------------------ rejection reporting
    def test_upload_reports_rejected_rows_in_summary(self, client, tokens):
        acc = _create_account(client, tokens["admin"])
        rows_data = [
            ["有效文章",  "2026/5/1",  "19:00", 100, 3, 5, 0, 0, 0, 0, 100, 1.0, 0.4, 0, 0, "Y"],
            ["",          "2026/5/1",  "19:00",  10, 0, 0, 0, 0, 0, 0,  10, 0.5, 0.1, 0, 0, "Y"],
            ["日期错误",  "INVALID",   "19:00",  20, 1, 2, 0, 0, 0, 0,  20, 0.8, 0.2, 0, 0, "Y"],
        ]
        xlsx_bytes = _make_wechat_xlsx_bytes(rows=rows_data)
        r = client.post(
            "/media/upload",
            files={"file": ("export.xlsx", xlsx_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"account_id": str(acc["id"])},
            headers=_auth(tokens["admin"]),
        )
        assert r.status_code == 200
        body = r.json()
        assert body["posts_upserted"] == 1
        assert body["rejected"] == 2


# ═══════════════════════════════════════════════════════════════════════════════
# Section 5 — Upload history  GET /media/uploads
# ═══════════════════════════════════════════════════════════════════════════════

class TestMediaUploadHistory:

    def _upload(self, client, tokens, acc_id: int, filename: str = "export.xlsx"):
        return client.post(
            "/media/upload",
            files={"file": (filename, _make_wechat_xlsx_bytes(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"account_id": str(acc_id)},
            headers=_auth(tokens["admin"]),
        )

    def test_history_requires_auth(self, client):
        r = client.get("/media/uploads")
        assert r.status_code in (401, 403)

    def test_viewer_cannot_see_history(self, client, tokens):
        r = client.get("/media/uploads", headers=_auth(tokens["viewer"]))
        assert r.status_code == 403

    def test_analyst_can_see_history(self, client, tokens):
        r = client.get("/media/uploads", headers=_auth(tokens["analyst"]))
        assert r.status_code == 200
        assert isinstance(r.json(), list)

    def test_history_shows_manual_uploads_in_reverse_order(self, client, tokens):
        acc = _create_account(client, tokens["admin"])
        self._upload(client, tokens, acc["id"], "first.xlsx")
        self._upload(client, tokens, acc["id"], "second.xlsx")

        r = client.get("/media/uploads", headers=_auth(tokens["analyst"]))
        runs = r.json()
        # Most recent first
        filenames = [run["filename"] for run in runs]
        assert filenames.index("second.xlsx") < filenames.index("first.xlsx")

    def test_history_excludes_api_sync_runs(self, client, tokens, monkeypatch):
        """GET /media/uploads must only return source='manual' runs."""
        acc = _create_account(client, tokens["admin"])
        # Trigger an API sync (which writes source='api' run)
        import app.views.media.routes as media_view

        class FakeClient:
            def __init__(self, *a, **kw): pass
            def fetch_article_total_rows(self, *a, **kw): return []

        monkeypatch.setattr(media_view, "WeChatOfficialClient", FakeClient)
        # We need app_id set for the account so sync doesn't fail
        from sqlalchemy import create_engine, text
        # Set app_id directly via DB so sync path can find credentials
        # (shortcut: just call the existing sync endpoint with env credentials)
        monkeypatch.setenv("WECHAT_OFFICIAL_APP_ID", "wx-test")
        monkeypatch.setenv("WECHAT_OFFICIAL_APP_SECRET", "sec")
        client.post(
            "/media/wechat/sync",
            json={"start_date": "2026-05-01", "end_date": "2026-05-01"},
            headers=_auth(tokens["admin"]),
        )

        # Now do a manual upload
        self._upload(client, tokens, acc["id"], "manual.xlsx")

        r = client.get("/media/uploads", headers=_auth(tokens["analyst"]))
        runs = r.json()
        # Must contain the manual upload
        assert any(run["filename"] == "manual.xlsx" for run in runs)
        # Must NOT contain API sync runs
        assert all(run["source"] == "manual" for run in runs)

    def test_history_entry_has_expected_fields(self, client, tokens):
        acc = _create_account(client, tokens["admin"])
        self._upload(client, tokens, acc["id"], "check.xlsx")

        r = client.get("/media/uploads", headers=_auth(tokens["analyst"]))
        run = r.json()[0]
        assert "id" in run
        assert "filename" in run
        assert "account_name" in run
        assert "posts_upserted" in run
        assert "metrics_upserted" in run
        assert "rejected" in run
        assert "started_at" in run
        assert run["source"] == "manual"
